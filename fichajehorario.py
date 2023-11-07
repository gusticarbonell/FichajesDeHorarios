import datetime
import openpyxl
from openpyxl import Workbook

# Diccionario para almacenar los registros de horario de cada usuario
registros = {}
usuario = input("Introduce tu nombre de usuario: ")

def iniciar_jornada():
    fecha_actual = datetime.date.today()
    if fecha_actual not in registros:
        registros[fecha_actual] = {}
    
    if usuario not in registros[fecha_actual]:
        registros[fecha_actual][usuario] = {'hora_inicio': datetime.datetime.now(), 'horas_trabajadas': datetime.timedelta()}
        print("Jornada iniciada.")
    else:
        print("Ya has iniciado una jornada hoy. Puedes pausar o finalizar la jornada actual.")

def pausar_jornada():
    fecha_actual = datetime.date.today()
    if fecha_actual in registros and usuario in registros[fecha_actual]:
        if 'pausa' not in registros[fecha_actual][usuario]:
            registros[fecha_actual][usuario]['pausa'] = datetime.datetime.now()
            print("Jornada pausada.")
        else:
            print("Ya has pausado la jornada. Puedes reanudarla.")

def reanudar_jornada():
    fecha_actual = datetime.date.today()
    if fecha_actual in registros and usuario in registros[fecha_actual] and 'pausa' in registros[fecha_actual][usuario]:
        pausa = registros[fecha_actual][usuario]['pausa']
        tiempo_pausado = datetime.datetime.now() - pausa
        # Restaurar la hora de inicio para que el tiempo no avance durante la pausa
        registros[fecha_actual][usuario]['hora_inicio'] = registros[fecha_actual][usuario]['hora_inicio'] + tiempo_pausado
        del registros[fecha_actual][usuario]['pausa']
        print("Jornada reanudada.")
    else:
        print("No hay una pausa registrada para reanudar.")


def terminar_jornada(workbook):
    fecha_actual = datetime.date.today()
    if fecha_actual in registros and usuario in registros[fecha_actual]:
        hora_inicio = registros[fecha_actual][usuario]['hora_inicio']
        horas_trabajadas = datetime.datetime.now() - hora_inicio
        registros[fecha_actual][usuario]['horas_trabajadas'] += horas_trabajadas

        horas = int(registros[fecha_actual][usuario]['horas_trabajadas'].total_seconds() // 3600)
        minutos = int((registros[fecha_actual][usuario]['horas_trabajadas'].total_seconds() % 3600) // 60)
        segundos = int(registros[fecha_actual][usuario]['horas_trabajadas'].total_seconds() % 60)
        print(f"Terminaste la jornada con {horas} horas, {minutos} minutos y {segundos} segundos trabajados el {fecha_actual}.")

        del registros[fecha_actual][usuario]

        # Guardar el registro en el archivo Excel
        guardar_registro_en_excel(fecha_actual, horas, minutos, segundos, workbook)

def guardar_registro_en_excel(fecha, horas, minutos, segundos, workbook):
    mes = fecha.strftime("%B")
    sheet_name = mes
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name)
        worksheet = workbook[sheet_name]
        worksheet.append(["Usuario", "Fecha", "Horas", "Minutos", "Segundos"])

    worksheet = workbook[sheet_name]
    worksheet.append([usuario, fecha, horas, minutos, segundos])

# Crear un archivo Excel
workbook = Workbook()

while True:
    print("Opciones:")
    print("1. Iniciar jornada")
    print("2. Pausar jornada")
    print("3. Reanudar jornada")
    print("4. Terminar jornada")
    print("5. Guardar en Excel")
    print("6. Salir")
    
    opcion = input("Elige una opción: ")
    
    if opcion == '1':
        iniciar_jornada()
        print()
    elif opcion == '2':
        pausar_jornada()
        print()
    elif opcion == '3':
        reanudar_jornada()
        print()
    elif opcion == '4':
        terminar_jornada(workbook)
        print()
    elif opcion == '5':
        excel_filename = "registros_horario.xlsx"
        workbook.save(excel_filename)
        print(f"Registros guardados en el archivo '{excel_filename}'.")
        if 'horas_trabajadas' in registros.get(datetime.date.today(), {}).get(usuario, {}):
            tiempo_trabajado = registros[datetime.date.today()][usuario]['horas_trabajadas']
            horas = int(tiempo_trabajado.total_seconds() // 3600)
            minutos = int((tiempo_trabajado.total_seconds() % 3600) // 60)
            segundos = int(tiempo_trabajado.total_seconds() % 60)
            print(f"Horas Trabajadas: {horas} horas, {minutos} minutos y {segundos} segundos.")
        print()
    elif opcion == '6':
        break
    else:
        print("Opción no válida. Inténtalo de nuevo.")
        print()
